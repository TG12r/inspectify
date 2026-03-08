import io
import openpyxl
from django.http import HttpResponse
from .forms_import import JobImportForm
# Vista para importar empleos desde Excel
from django.contrib.auth.decorators import login_required
@login_required
def job_import(request):
    if not (request.user.is_superuser or request.user.is_staff or request.user.is_editor()):
        return redirect('job_list')
    if request.method == 'POST':
        form = JobImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                title, company, location, description, salary_range, url, apply_link, source, posted_at, is_active = row[:10]
                if not title or not company or not url:
                    continue
                from .models import JobOffer
                JobOffer.objects.create(
                    title=title,
                    company=company,
                    location=location or '',
                    description=description or '',
                    salary_range=salary_range or '',
                    url=url,
                    apply_link=apply_link or '',
                    source=source or 'Unknown',
                    posted_at=posted_at,
                    is_active=bool(is_active),
                )
                created += 1
            return render(request, 'jobs/job_import_result.html', {'created': created})
    else:
        form = JobImportForm()
    return render(request, 'jobs/job_import.html', {'form': form})

# Vista para descargar formato de ejemplo
@login_required
def job_import_format(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        'title', 'company', 'location', 'description', 'salary_range', 'url', 'apply_link', 'source', 'posted_at', 'is_active'
    ]
    ws.append(headers)
    example = [
        'Inspector API 510', 'Empresa S.A.', 'Ciudad', 'Inspección de tanques según API 510. Experiencia requerida.',
        'USD 2000-3000/mes', 'https://ejemplo.com/oferta', 'https://ejemplo.com/apply', 'LinkedIn', '2026-03-08', 1
    ]
    ws.append(example)

    # Mejorar formato visual
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        max_length = 18
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    for cell in ws[2]:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_empleos.xlsx"'
    return response
from django.shortcuts import render, get_object_or_404, redirect
from .models import JobOffer
from .forms import JobOfferForm
from django.contrib.auth.decorators import login_required
from core.models import User
def user_can_edit_job(user, job):
    return user.is_authenticated and (user.is_superuser or user.is_staff or user.is_editor()) and job.author == user

# Solo editores y admins pueden crear
@login_required
def job_create(request):
    if not (request.user.is_superuser or request.user.is_staff or request.user.is_editor()):
        return redirect('job_list')
    if request.method == 'POST':
        form = JobOfferForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.author = request.user
            job.save()
            return redirect('job_detail', pk=job.pk)
    else:
        form = JobOfferForm()
    return render(request, 'jobs/job_form.html', {'form': form})

# Solo el autor (editor) o admin puede editar
@login_required
def job_edit(request, pk):
    job = get_object_or_404(JobOffer, pk=pk)
    if not user_can_edit_job(request.user, job):
        return redirect('job_detail', pk=pk)
    if request.method == 'POST':
        form = JobOfferForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            return redirect('job_detail', pk=job.pk)
    else:
        form = JobOfferForm(instance=job)
    return render(request, 'jobs/job_form.html', {'form': form, 'job': job})
from django.core.paginator import Paginator
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.core.management import call_command
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q

@staff_member_required
def trigger_scraping(request):
    if request.method == 'POST':
        try:
            # Get params from modal form
            keywords = request.POST.get('keywords', 'API 510')
            source = request.POST.get('source', 'all')
            location = request.POST.get('location', '')
            limit = int(request.POST.get('limit', 10))
            
            # Handle 'TODAS' keyword logic (same as in tasks.py)
            if keywords.upper() in ['TODAS', 'ALL']:
                default_keywords = ['API 510', 'API 570', 'API 653', 'CWI', 'NDT', 'Welding Inspector', 'QA/QC Inspector']
                for kw in default_keywords:
                    call_command('scrape_jobs', source=source, keywords=kw, location=location, limit=limit)
            else:
                call_command('scrape_jobs', source=source, keywords=keywords, location=location, limit=limit)
                
            messages.success(request, f'Scraping manual ejecutado: {keywords} en {source}')
        except Exception as e:
            messages.error(request, f'Error al ejecutar scraping: {str(e)}')
            
        return redirect('admin:index')
    
    return redirect('admin:index')

@staff_member_required
def schedule_scraping(request):
    if request.method == 'POST':
        try:
            # Get params from modal form
            keywords = request.POST.get('keywords', 'API 510')
            source = request.POST.get('source', 'all')
            location = request.POST.get('location', '')
            limit = int(request.POST.get('limit', 10))
            frequency = request.POST.get('frequency', 'daily')
            
            # Calculate repeat interval
            repeat = None
            if frequency == 'daily':
                repeat = 24 * 60 * 60
            elif frequency == 'weekly':
                repeat = 7 * 24 * 60 * 60
            elif frequency == 'monthly':
                repeat = 30 * 24 * 60 * 60
            # 'once' implies repeat=None (default)

            # Schedule daily task
            from jobs.tasks import scrape_all_jobs_task
            
            # We allow multiple schedules with different params. 
            # To prevent exact duplicates, we could check args, but for now let's just schedule it.
            # The verbose_name or checking specific args in Task model is complex with hashing.
            # We will just schedule it. User can manage duplicates in Django Admin > Background Tasks if needed.
            
            scrape_all_jobs_task(
                keywords=keywords, 
                source=source, 
                location=location, 
                limit=limit, 
                repeat=repeat
            )
            
            freq_label = {
                'daily': 'Diariamente',
                'weekly': 'Semanalmente',
                'monthly': 'Mensualmente',
                'once': 'Una vez'
            }.get(frequency, 'Diariamente')
            
            messages.success(request, f'Scraping programado: {keywords} en {source} ({freq_label})')
            
        except Exception as e:
            messages.error(request, f'Error al programar scraping: {str(e)}')
            
        return redirect('admin:index')
    
    return redirect('admin:index')

def job_list(request):
    jobs = JobOffer.objects.filter(is_active=True)
    
    # Filter by search query
    query = request.GET.get('q')
    if query:
        jobs = jobs.filter(title__icontains=query) | jobs.filter(company__icontains=query) | jobs.filter(description__icontains=query)
    
    # Filter by location
    location = request.GET.get('location')
    if location:
        jobs = jobs.filter(location__icontains=location)

    # Filter by source
    source = request.GET.get('source')
    if source:
        jobs = jobs.filter(source__iexact=source)

    # Filter by remote
    is_remote = request.GET.get('remote')
    if is_remote == 'true' or is_remote == 'on':
        jobs = jobs.filter(remote=True)
    
    # Filter by last 24 hours
    hours = request.GET.get('hours')
    if hours:
        try:
            hours_int = int(hours)
            time_cutoff = timezone.now() - timedelta(hours=hours_int)
            # Use posted_at if available, otherwise use created_at
            jobs = jobs.filter(Q(posted_at__gte=time_cutoff.date()) | Q(created_at__gte=time_cutoff))
        except (ValueError, TypeError):
            pass
    
    # Filter by experience level
    level = request.GET.get('level')
    if level == 'junior':
        jobs = jobs.filter(Q(title__icontains='junior') | Q(title__icontains='entry') | 
                          Q(title__icontains='trainee') | Q(description__icontains='junior') |
                          Q(description__icontains='entry level'))
    elif level == 'senior':
        jobs = jobs.filter(Q(title__icontains='senior') | Q(title__icontains='lead') | 
                          Q(title__icontains='principal') | Q(title__icontains='manager') |
                          Q(description__icontains='senior'))
    
    # Filter by skill
    skill = request.GET.get('skill')
    if skill:
        jobs = jobs.filter(Q(title__icontains=skill) | Q(description__icontains=skill))

    # Ordering
    sort_by = request.GET.get('sort', '-posted_at')
    if sort_by in ['posted_at', '-posted_at', 'salary', '-salary']:
        jobs = jobs.order_by(sort_by)
    else:
        # Default ordering by created_at if posted_at is null
        jobs = jobs.order_by('-created_at')

    # Pagination
    paginator = Paginator(jobs, 10) # 10 jobs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'jobs': page_obj, # Pass page_obj instead of jobs
        'query': query,
        'location': location,
        'selected_source': source,
        'is_remote': is_remote,
        'sort_by': sort_by,
        'level': level,
        'skill': skill,
        'hours': hours,
        'sources': JobOffer.objects.values_list('source', flat=True).distinct().order_by('source')
    }
    return render(request, 'jobs/job_list.html', context)

def job_detail(request, pk):
    job = get_object_or_404(JobOffer, pk=pk)
    return render(request, 'jobs/job_detail.html', {'job': job})
